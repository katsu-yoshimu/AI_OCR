import argparse
import os
import re
from pathlib import Path
import google.generativeai as genai

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import mimetypes

def get_gemini_api_key():
    """環境変数からGemini APIキーを取得する"""
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("エラー: 環境変数 'GEMINI_API_KEY' が設定されていません。設定してください。")
        exit(1)

def get_mime_type(filepath):
    """ファイルの拡張子からMIMEタイプを取得する"""
    mime_type, _ = mimetypes.guess_type(filepath)
    if not mime_type:
        raise ValueError(f"対応していないファイル形式です: {filepath}")
    return mime_type

def read_file(filepath):
    """ファイルを読み込み、バイナリデータを返す"""
    mime_type = get_mime_type(filepath)
    if not mime_type:
      raise ValueError(f"対応していないファイル形式です: {filepath}")
    with open(filepath, "rb") as f:
        return f.read(), mime_type

def create_prompt(mime_type, outfiletype):
    """プロンプトを作成する"""
    file_type = "PDF" if mime_type == "application/pdf" else "画像"
    output_format = {
      "txt": "プレーンテキスト",
      "md": "マークダウン",
      "csv": "CSV",
      "html": "HTML",
      "xlsx": "マークダウン",
    }.get(outfiletype, "プレーンテキスト")
    prompt = f"この{file_type}を読み込んで{output_format}形式で出力してください。"
    if outfiletype == "html":
      prompt += "表はstyleを指定して見やすくしてください。"
    return prompt

def extract_file_content(response_text):
    """Geminiのレスポンスからコードブロック＝ファイル内容を抽出する"""
    # コードブロックの正規表現パターン
    pattern = r'```(?:[^\n]*\n)?([\s\S]*?)(?:```|$)'
    
    # 最初のコードブロックを検索
    match = re.search(pattern, response_text)
    
    if match:
        # コードブロックの内容を返す
        return match.group(1).strip()
    else:
        # コードブロックが見つからない場合はNoneを返す
        return None

def markdown_to_excel(markdown_text, excel_file):
    """マークダウン形式テキストからExcelファイルを生成する"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    row_num = 1
    table_data = []
    in_table = False
    
    for line in markdown_text.splitlines():
        line = line.strip()
        
        if line.startswith("#"):
            if in_table:
                output_table(sheet, table_data, row_num)
                row_num += len(table_data) + 1
                table_data = []
                in_table = False
            header_level = len(line.split(" ")[0])
            header_text = line[header_level+1:].strip()
            cell = sheet.cell(row=row_num, column=1, value=header_text)
            cell.font = Font(bold=True)
            row_num += 1
        elif line.startswith("|"):
            in_table = True
            row_data = [cell.strip() for cell in line.split("|")[1:-1]]
            row_data = [cell.replace("<br>", "\n").replace("<BR>", "\n") for cell in row_data]
            # row_data が すべて「-」の時＝表のヘッダとデータの区切り行は追加しない
            if (is_markdown_table_separator(row_data)) == False:
                table_data.append(row_data)
        else:
            if in_table:
                output_table(sheet, table_data, row_num)
                row_num += len(table_data) + 1
                table_data = []
                in_table = False
            if line:
               cell = sheet.cell(row=row_num, column=1, value=line)
               row_num += 1

    if in_table:
        output_table(sheet, table_data, row_num)

    adjust_column_width(sheet)
    wb.save(excel_file)
    
def output_table(sheet, table_data, start_row):
    """マークダウン形式テキストの表をExcelファイルを書きこむ"""
    if not table_data:
        return
        
    header = table_data[0]
    data = table_data[1:]
    
    header_row = start_row
    data_start_row = start_row + 1
    
    # ヘッダー出力
    for col, header_text in enumerate(header, start=2):
        cell = sheet.cell(row=header_row, column=col, value=header_text)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # データ出力
    for row, row_data in enumerate(data, start=data_start_row):
        for col, cell_value in enumerate(row_data, start=2):
            cell = sheet.cell(row=row, column=col, value=cell_value)
            cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
    
    # 罫線設定
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in range(start_row,start_row + len(table_data)):
        for col in range(2,len(header)+2):
            sheet.cell(row=row,column=col).border = border

def is_markdown_table_separator(strings):
    """マークダウン形式テキストの表のヘッダとデータの区切り行を判定する"""
    pattern = r'^(:?-+):?(:?-*)$'

    # すべての文字列が条件を満たすか判定
    return all(re.fullmatch(pattern, s) for s in strings)

def adjust_column_width(sheet):
    """Excelファイルの列の幅を調整する。B列以降、最大50。"""
    MAX_COLUMN_WIDTH = 50

    for col in range(2, sheet.max_column + 1):
        max_length = 0
        for cell in sheet[get_column_letter(col)]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > MAX_COLUMN_WIDTH:
            adjusted_width = MAX_COLUMN_WIDTH
        sheet.column_dimensions[get_column_letter(col)].width = adjusted_width

def process_response(response, outfiletype, outfile):
    # レスポンス処理とファイル出力のロジック
    response_text = response.text
    file_content = extract_file_content(response_text)
    if file_content == None:
        file_content = response_text

    if outfiletype == "xlsx":
        markdown_to_excel(file_content, outfile)
        with open(outfile + ".temp.md", "w", encoding="utf-8") as f:
            f.write(file_content)
    elif outfile:
        if outfiletype == "csv":
            import locale
            output_encoding = locale.getpreferredencoding()
        else:
            output_encoding = "utf-8"
        with open(outfile, "w", encoding=output_encoding) as f:
            f.write(file_content)
    else:
        print(file_content)


def main():
    """メイン処理"""
    parser = argparse.ArgumentParser(description="PDF/画像ファイルをOCRして出力します。")
    parser.add_argument("-infile", required=True, help="入力ファイル (PDFまたは画像)")
    parser.add_argument("-outfile", help="出力ファイル (省略時は標準出力)")
    parser.add_argument("-outfiletype", help="出力ファイル形式 (txt, md, csv, html, xlsx) (省略時は出力ファイルの拡張子)")
    args = parser.parse_args()

    infile = args.infile.lower()
    outfile = args.outfile
    outfiletype = args.outfiletype
    
    if outfile:
      outfile = outfile.lower()
      if not outfiletype:
         outfiletype = Path(outfile).suffix.lower()[1:]
    else:
      if not outfiletype:
        outfiletype = "txt"

    print(f"入力ファイル: {infile}")
    print(f"出力ファイル: {outfile if outfile else '標準出力'}")
    print(f"出力ファイル形式: {outfiletype}")

    if not outfile and outfiletype == "xlsx":
        print("エラー: 標準出力に xlsx 形式はサポートされていません。")
        exit(1)

    try:
        api_key = get_gemini_api_key()
        genai.configure(api_key=api_key)

        model = genai.GenerativeModel("gemini-2.0-flash-exp")
        file_data, mime_type = read_file(infile)

        prompt = create_prompt(mime_type, outfiletype)
        print(f"Geminiプロンプト: {prompt}")
        
        content = [{'mime_type': mime_type, 'data': file_data}, prompt]
        response = model.generate_content(content)
        print("Geminiレスポンス:")
        print(response)

        process_response(response, outfiletype, outfile)

    except FileNotFoundError as e:
        print(f"ファイルが見つかりません: {e}")
    except ValueError as e:
        print(f"無効な入力です: {e}")
    except Exception as e:
        print(f"不明なエラーが発生しました: {e}")


if __name__ == "__main__":
    main()
